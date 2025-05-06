import pandas as pd
import math
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# إعدادات
input_file = "omni-mpsetupbymatch-en-all.xlsx"
sheet_name = "MP Item Setup By Match"
upc_column_index = 5
output_folder = "split_files"
max_total_rows = 9980
header_row_count = 6
real_header_row = 4  # الصف 5 في الإكسل = index 4 في بايثون
max_data_rows = max_total_rows - header_row_count

# قراءة الملف
df = pd.read_excel(input_file, sheet_name=sheet_name, header=None, dtype={upc_column_index: str})
header_rows = df.iloc[:header_row_count]
data_rows = df.iloc[header_row_count:]
filtered_data = data_rows[data_rows[upc_column_index].astype(str).str.len() >= 12]
num_parts = math.ceil(len(filtered_data) / max_data_rows)
os.makedirs(output_folder, exist_ok=True)

# تنسيقات
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# حفظ وتقسيم
for i in range(num_parts):
    start = i * max_data_rows
    end = start + max_data_rows
    chunk = pd.concat([header_rows, filtered_data.iloc[start:end]], ignore_index=True)

    filename = f"products_part_{i+1} - الجزء_{i+1}.xlsx"
    filepath = os.path.join(output_folder, filename)
    chunk.to_excel(filepath, index=False, header=False)

    wb = load_workbook(filepath)
    ws = wb.active

    for row in ws.iter_rows(min_row=real_header_row+1, max_row=header_row_count):
        for cell in row:
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

    wb.save(filepath)

print(f"✅ الملفات تم تقسيمها وتنسيق رؤوس الأعمدة بنجاح في المجلد '{output_folder}'")
